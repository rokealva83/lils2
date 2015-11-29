# -- coding: utf-8 --
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.writer.excel import save_virtual_workbook
import barcode
from barcode.writer import ImageWriter
from numpy import array


class Exporter(object):

    def export(self, model):
        raise NotImplementedError()


class BoxExporter(Exporter):
    metadata_font = Font(
        size=16,
        bold=True
    )

    metadata_style = Alignment(
        vertical='center',
        horizontal='center',
        shrink_to_fit=True,
        shrinkToFit=False
    )

    table_header_font = Font(
        bold=True
    )

    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def _write_customer_title(self, work_sheet, customer):
        work_sheet.append((
            'Customer', str(customer),
        ))

        row = self._get_last_row(work_sheet)
        row.font = self.metadata_font
        row.alignment = self.metadata_style

        if customer.invoice:
            work_sheet.append((
                'Invoice', customer.invoice,
            ))

            row = self._get_last_row(work_sheet)
            row.font = self.metadata_font
            row.alignment = self.metadata_style
        else:
            self._write_empty_line(work_sheet)


    def _write_box_title(self, work_sheet, box):
        work_sheet.append((
            'Box', str(box),
        ))

        row = self._get_last_row(work_sheet)
        row.font = self.metadata_font
        row.alignment = self.metadata_style

    def _write_customer_total_weight(self, work_sheet, customer):
        if customer.total_weight:
            work_sheet.append((
                'Total weight', customer.total_weight
            ))

            row = self._get_last_row(work_sheet)
            row.font = self.metadata_font
            row.alignment = self.metadata_style

    def _write_product_table_header(self, work_sheet):
        work_sheet.append((
            'Barcode',
            'Name',
            'Order',
            'Quantity',
        ))

        row = self._get_last_row(work_sheet)
        row.font = self.table_header_font
        row.alignment = self.metadata_style

    def _write_empty_line(self, work_sheet):
        work_sheet.append([])

    def _get_last_row(self, work_sheet):
        rows = work_sheet.row_dimensions
        return rows[max(rows.keys())]

    def _write_product_table_row(self, work_book, work_sheet, product, line, percent_top):
        writer = ImageWriter()
        writer.set_options({'font_size': 18, 'text_distance': 2, 'quiet_zone': 2})
        ean = barcode.get('ean13', product.barcode, writer=writer)
        file_name = ean.save('barcode', {'quiet_zone': 2, 'text_distance': 2, 'module_height': 8})
        size = openpyxl.drawing.Image(file_name).image.size * array([0.35, 0.35])
        img = openpyxl.drawing.Image(file_name, size=size)        
        line_number = 'A%s' % line
        img.anchor(work_book.worksheets[0].cell(line_number))
        img.drawing.top = percent_top

        work_sheet.append((
            product.barcode,
            product.name,
            product.order,
            product.quantity,
        ))
        row = self._get_last_row(work_sheet)
        row.alignment = self.metadata_style
        work_sheet.append((work_sheet.add_image(img),)),
        return img.drawing.top

    def export(self, box):
        work_book = Workbook()

        work_sheet = work_book.active
        work_sheet.title = 'Box'

        self._write_customer_title(work_sheet, box.parent)
        self._write_box_title(work_sheet, box)
        self._write_empty_line(work_sheet)
        self._write_product_table_header(work_sheet)
        work_sheet = column_dement(work_sheet)
        work_sheet.row_dimensions[1].height = 30
        work_sheet.row_dimensions[2].height = 30
        work_sheet.row_dimensions[3].height = 30
        line = 6
        percent_top = 172
        for product in box.productpurchase_set.all():
            self._write_product_table_row(work_book, work_sheet, product, line, percent_top)
            work_sheet.row_dimensions[line].height = 60
            line = line + 2
            percent_top = percent_top + 100

        return save_virtual_workbook(work_book)


class CustomerExporter(BoxExporter):

    def export(self, customer):
        work_book = Workbook()
        work_sheet = work_book.active
        work_sheet.title = 'Boxs'
        self._write_customer_title(work_sheet, customer)
        self._write_customer_total_weight(work_sheet, customer)
        work_sheet.row_dimensions[1].height = 30
        work_sheet.row_dimensions[2].height = 30
        self._write_empty_line(work_sheet)
        line = 4
        percent_top = 236
        num_box = 1
        for box in customer.box_set.all():
            self._write_box_title(work_sheet, box)
            self._write_product_table_header(work_sheet)
            work_sheet.row_dimensions[3].height = 30
            work_sheet.row_dimensions[4].height = 30
            work_sheet.row_dimensions[line+1].height = 30
            work_sheet = column_dement(work_sheet)
            line += 3

            for product in box.productpurchase_set.all():
                top = self._write_product_table_row(work_book, work_sheet, product, line, percent_top)
                work_sheet.row_dimensions[line].height = 60
                line  = line+2
                percent_top = percent_top + 100
            self._write_empty_line(work_sheet)
            line  = line+1
            percent_top = percent_top + 100
            num_box = num_box + 1
        return save_virtual_workbook(work_book)

def column_dement(work_sheet):
    work_sheet.column_dimensions['A'].width = 25
    work_sheet.column_dimensions['B'].width = 17
    work_sheet.column_dimensions['C'].width = 19
    work_sheet.column_dimensions['D'].width = 8
    return work_sheet

from .models import Box, Customer

__exporters_instances = {
    Box: BoxExporter(),
    Customer: CustomerExporter(),
}

def get_exporter(model):
    return __exporters_instances[model]
